import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from participacio import variables, reglas, variable_siempre_presente
from excepciones import excepciones

# Cargar el archivo .sav
df, meta = pyreadstat.read_sav('snPART_2024.sav')

responsables_df = pd.read_excel('responsables.xlsx')

arxiu = f'incongruencias_por_regla_tot.xlsx'
# Realizar un left join con la columna 'q6' (en común) y añadir la columna 'Responsable' de responsables.xlsx
df = df.merge(responsables_df[['q6', 'Responsable']], how='left', on='q6')

# Filtrar el DataFrame excluyendo las excepciones por 'q6'
df = df[~df['q6'].astype(str).isin([ex['q6'] for ex in excepciones.values()])]

# Inicializar Excel writer
with pd.ExcelWriter(arxiu, engine='openpyxl') as writer:

    # Crear un DataFrame inicial para la hoja de resumen de responsables, incluyendo las columnas q7 y q9
    resumen_responsables = df[['q6', 'Responsable', 'q7', 'q9']].drop_duplicates().set_index('q6')

    # Aplicar las reglas y escribir las hojas correspondientes
    for key, regla in reglas.items():
        # Evaluar la condición
        resultado = eval(regla['condicion'])
        
        # Si hay registros que cumplan la condición de error, marcarlos
        if resultado.any():
            # Filtrar los casos que cumplen con la incongruencia
            incongruencias_encontradas = df[resultado].copy()
            incongruencias_encontradas['error'] = regla['mensaje']
            
            # Calcular las columnas adicionales según la regla específica
            if key == 'organs_suma_incorrecte':
                incongruencias_encontradas['total calculat'] = df[variables['suma_organos_1']].sum(axis=1, skipna=True)
            if key == 'processos_planes_incorrectes':
                incongruencias_encontradas['total calculat'] = df[variables['procesos_marcados_a_planes']].sum(axis=1, skipna=True)
            if key == 'treballadors_hores_incorrectes':
                incongruencias_encontradas['total calculat'] = df[variables['trabajadores_horas_total']].fillna(0).sum(axis=1)
            if key == 'fondos_participacio_incorrectes':
                incongruencias_encontradas['total calculat'] = df[variables['fondos_participacion']].fillna(0).sum(axis=1)
            if key == 'equipaments_suma_zero':
                incongruencias_encontradas['total calculat'] = df[variables['total_equipamiento']].fillna(0).sum(axis=1)
            
            # Seleccionar las variables relacionadas con la regla actual y añadir las variables siempre presentes
            columnas_a_imprimir = ['Responsable'] + variable_siempre_presente + regla['variables'] + ['error']

            # Ordenar por población descendiente (q7) y luego por Responsable de forma alfabética
            incongruencias_encontradas = incongruencias_encontradas.sort_values(by=['Responsable', 'q7'], ascending=[True, False])
            
            # Escribir las incongruencias en una nueva hoja del archivo Excel
            incongruencias_encontradas[columnas_a_imprimir].to_excel(writer, sheet_name=key, index=False)

            
            # Marcar los municipios que están presentes en la regla actual
            resumen_responsables.loc[incongruencias_encontradas['q6'], key] = "Sí"

    # Rellenar los valores faltantes en la tabla resumen con "No"
    resumen_responsables.fillna("No", inplace=True)

    # Filtrar para incluir solo municipios con al menos un "Sí"
    resumen_responsables = resumen_responsables.loc[resumen_responsables.eq("Sí").any(axis=1)]

    resumen_responsables = resumen_responsables.sort_values(by=['Responsable', 'q7'], ascending=[True, False])

    # Escribir la hoja de resumen en primer lugar
    resumen_responsables.to_excel(writer, sheet_name='Resum per municipi')

# Crear hipervínculos en la hoja de resumen
wb = load_workbook(arxiu)
ws_resum = wb['Resum per municipi']

wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index('Resum per municipi')))

# Añadir los hipervínculos
# Iterar sobre las filas de la hoja de resumen de responsables
for row in range(2, ws_resum.max_row + 1):
    # Iterar sobre las columnas que contienen el "Sí" o "No" para las reglas
    for col in range(4, ws_resum.max_column + 1):
        # Verificar si el valor de la celda es "Sí"
        if ws_resum.cell(row=row, column=col).value == "Sí":
            # Obtener el nombre de la hoja de la regla desde la primera fila
            hoja_regla = ws_resum.cell(1, col).value
            
            # Buscar el valor de la columna q6 en la fila actual
            q6_value = ws_resum.cell(row, 1).value
            
            # Abrir la hoja correspondiente
            ws_regla = wb[hoja_regla]

            # Buscar la columna que contiene q6 en la hoja de la regla
            col_q6 = None
            for col_num in range(1, ws_regla.max_column + 1):
                if ws_regla.cell(row=1, column=col_num).value == 'q6':
                    col_q6 = col_num
                    break
            
            # Si la columna q6 existe en la hoja de la regla
            if col_q6 is not None:
                # Buscar la fila que contiene el valor q6_value en la columna q6
                row_q6 = None
                for row_num in range(2, ws_regla.max_row + 1):
                    if ws_regla.cell(row=row_num, column=col_q6).value == q6_value:
                        row_q6 = row_num
                        break
                
                # Si se encontró la fila con el valor de q6
                if row_q6 is not None:
                    # Crear el hipervínculo hacia la celda correcta en la hoja de la regla
                    link = f"#'{hoja_regla}'!A{row_q6}"
                    ws_resum.cell(row, col).hyperlink = link
                    ws_resum.cell(row, col).style = "Hyperlink"


# Guardar los cambios
wb.save(arxiu)

print("Las incongruencias se han guardado en el archivo",arxiu)
